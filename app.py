from flask import Flask, request, render_template
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side
import os

app = Flask(__name__)

# Ruta para cargar la página HTML
@app.route('/')
def index():
    return render_template('formulario.html')

# Ruta para manejar la solicitud POST al enviar el formulario
@app.route('/submit', methods=['POST'])
def submit():
    try:
        # Datos del cliente
        nombre = request.form['nombre']
        dni = request.form['dni']
        telefono = request.form['telefono']
        email = request.form['email']
        direccion = request.form['direccion']
        distrito = request.form['distrito']

        # Datos del pedido
        tienda = request.form['tienda']
        fecha_pedido = request.form['fecha-pedido']
        descripcion = request.form['descripcion']
        peso = request.form['peso']
        alto = request.form['alto']
        bajo = request.form['bajo']
        ancho = request.form['ancho']
        direccion_recogida = request.form['direccion-recogida']
        direccion_entrega = request.form['direccion-entrega']
        fecha_entrega = request.form['fecha-entrega']

        # Guardar la imagen subida en una carpeta
        imagen = request.files['imagen-producto']
        imagen_path = None
        if imagen:
            imagen_path = os.path.join('imagenes_productos', imagen.filename)
            imagen.save(imagen_path)

        # Abre el archivo Excel o crea uno si no existe
        if os.path.exists('registros_clientes_pedidos.xlsx'):
            workbook = load_workbook('registros_clientes_pedidos.xlsx')
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            # Cabeceras de columnas
            sheet.append(['Nombre', 'DNI', 'Teléfono', 'Correo Electrónico', 'Dirección', 'Distrito',
                          'Nombre de Tienda', 'Fecha del Pedido', 'Descripción', 'Peso (g)', 'Dimensiones (cm)',
                          'Dirección de Recogida', 'Dirección de Entrega', 'Fecha de Entrega', 'Imagen del Producto'])

        # Agregar los datos a una nueva fila
        fila = sheet.max_row + 1
        sheet.append([nombre, dni, telefono, email, direccion, distrito,
                      tienda, fecha_pedido, descripcion, peso, f'{alto}x{bajo}x{ancho}',
                      direccion_recogida, direccion_entrega, fecha_entrega])

        # Insertar la imagen en la última fila, si existe
        if imagen_path:
            img = Image(imagen_path)
            img.width = 100  # Ajustar el ancho de la imagen
            img.height = 100  # Ajustar la altura de la imagen
            celda_imagen = f'O{fila}'  # Columna O para la imagen
            sheet.add_image(img, celda_imagen)

            # Ajustar el tamaño de la celda para que la imagen quepa bien
            sheet.row_dimensions[fila].height = 80  # Altura de la fila
            sheet.column_dimensions['O'].width = 18  # Ancho de la columna O

        # Alinear y bordear las celdas para que todo quede ordenado
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        for col in range(1, 16):
            cell = sheet.cell(row=fila, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Guarda el archivo Excel
        workbook.save('registros_clientes_pedidos.xlsx')

        return 'Datos registrados correctamente'
    
    except Exception as e:
        return f'Error al procesar los datos: {str(e)}'

if __name__ == '__main__':
    # Crear la carpeta para guardar imágenes si no existe
    if not os.path.exists('imagenes_productos'):
        os.makedirs('imagenes_productos')
    app.run(debug=True)
